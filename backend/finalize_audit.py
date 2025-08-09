import os
import shutil
import json
import pandas as pd
from fastapi import APIRouter, Form
from fastapi.responses import JSONResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

router = APIRouter()

RESULTS_DIR = "results"
WHO_DATA_FILE = "data/WhoData.csv"

@router.post("/finalize_audit")
async def finalize_audit(
    full_vin: str = Form(...),
    total_ok: int = Form(...),
    total_notok: int = Form(...),
    total_pending: int = Form(...),
    component_statuses: str = Form(...)
):
    try:
        # âœ… 1. Rename Folder from (Ongoing) â†’ (Done)
        ongoing_folder = os.path.join(RESULTS_DIR, f"{full_vin} (Ongoing)")
        done_folder = os.path.join(RESULTS_DIR, f"{full_vin} (Done)")

        if os.path.exists(ongoing_folder):
            shutil.move(ongoing_folder, done_folder)
        elif not os.path.exists(done_folder):
            return JSONResponse({"status": "error", "message": "No ongoing folder found"}, status_code=400)

        # âœ… 2. Update WhoData CSV
        if not os.path.exists(WHO_DATA_FILE):
            return JSONResponse({"status": "error", "message": "WhoData.csv not found"}, status_code=404)

        df = pd.read_csv(WHO_DATA_FILE)
        df.columns = [c.strip() for c in df.columns]  # normalize column names

        # Get current time for both CSV and summary
        current_time = datetime.now()

        # Find the row with the VIN and update its status (5th column: index 4) and date (6th column: index 5)
        row_idx = df[df.iloc[:, 0].astype(str).str.strip() == full_vin.strip()].index

        if not row_idx.empty:
            status = (
                "Incomplete" if total_pending > 0 else
                "Finished (OK)" if total_notok == 0 else
                "Finished (NOT OK)"
            )
            audit_date = current_time.strftime("%Y-%m-%d")
            
            df.iat[row_idx[0], 4] = status      # Update Status column
            df.iat[row_idx[0], 5] = audit_date  # Update AuditDate column
            df.to_csv(WHO_DATA_FILE, index=False)
            
            # Get person details for Excel export
            person_name = df.iat[row_idx[0], 3]  # PersonName column (index 3)
            person_pno = df.iat[row_idx[0], 2]   # PersonPno column (index 2)
        else:
            return JSONResponse({"status": "error", "message": f"{full_vin} not found in WhoData"}, status_code=404)

        # âœ… 3. Write Final Summary
        component_data = json.loads(component_statuses)
        timestamp = current_time.strftime("%Y-%m-%d %H:%M:%S")
        day_name = current_time.strftime("%A")

        final_verdict_text = (
            "INCOMPLETE" if total_pending > 0 else
            "OK" if total_notok == 0 else
            "NOT OK"
        )

        summary_lines = [
            "Final Audit Summary",
            "===================",
            f"Audit Completed: {timestamp}",
            f"Day: {day_name}",
            "",
            f"Full VIN: {full_vin}",
            f"Total OK: {total_ok}",
            f"Total NOT OK: {total_notok}",
            f"Total Pending: {total_pending}",
            f"Final Verdict: {final_verdict_text}",
            "",
            "Detailed Component Status:",
            "========================="
        ]

        for category, components in component_data.items():
            summary_lines.append(f"\n[{category.upper()}]")
            for component_name, status in components.items():
                summary_lines.append(f"  - {component_name}: {status.upper()}")

        summary_file_path = os.path.join(done_folder, "FinalSummary.txt")
        with open(summary_file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(summary_lines))

        # âœ… 4. Create Excel Summary Report
        excel_file_path = os.path.join(done_folder, "FinalAuditReport.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Summary"

        # Header styling
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal="center")

        # Excel Headers
        headers = ["Field", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align

        # Excel Data
        excel_data = [
            ["Audit Completed", timestamp],
            ["Day", day_name],
            ["Full VIN", full_vin],
            ["Person Name", person_name],
            ["Person ID", person_pno],
            ["Total OK", total_ok],
            ["Total NOT OK", total_notok],
            ["Total Pending", total_pending],
            ["Final Verdict", final_verdict_text]
        ]

        # Add basic info
        for row, (field, value) in enumerate(excel_data, 2):
            ws.cell(row=row, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        # Add component details
        current_row = len(excel_data) + 3
        ws.cell(row=current_row, column=1, value="Component Details").font = Font(bold=True, size=11)
        current_row += 1

        for category, components in component_data.items():
            ws.cell(row=current_row, column=1, value=f"[{category.upper()}]").font = Font(bold=True)
            current_row += 1
            for component_name, status in components.items():
                ws.cell(row=current_row, column=1, value=f"  {component_name}")
                ws.cell(row=current_row, column=2, value=status.upper())
                current_row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(excel_file_path)

        return JSONResponse({
            "status": "success",
            "message": "Audit finalized successfully",
            "final_verdict": final_verdict_text
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)